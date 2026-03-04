"""
Multi-Site Price Tracker
Track the same EAN product across multiple garden/retail sites
and find the best price.

Author: Mohamed Amine
Stack: requests, BeautifulSoup, pandas, openpyxl
Sites: intratuin.nl, gamma.nl, hornbach.nl (extendable)
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
from datetime import datetime
import time
import json
from pathlib import Path


HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "nl-NL,nl;q=0.9",
}

# ---------------------------------------------------------------------------
# Site-specific scrapers — each returns {"price": float|None, "name": str|None,
#                                         "url": str|None, "in_stock": bool}
# ---------------------------------------------------------------------------

SITES = {
    "intratuin.nl": {
        "search_url": "https://www.intratuin.nl/search?q={ean}",
        "product_link_sel": "a.product-tile__link, a[href*='/p/']",
        "price_sel": "span.product-price__current, span[class*='price--current']",
        "name_sel": "h1.product-title, h1[class*='title']",
        "stock_sel": "span[class*='stock']",
    },
    "gamma.nl": {
        "search_url": "https://www.gamma.nl/assortiment/zoeken?text={ean}",
        "product_link_sel": "a.product-tile, a[href*='/assortiment/p/']",
        "price_sel": "[class*='price__current'], [data-testid='price']",
        "name_sel": "h1[class*='title'], h1[class*='name']",
        "stock_sel": "[class*='stock'], [class*='availability']",
    },
    "hornbach.nl": {
        "search_url": "https://www.hornbach.nl/zoeken/?q={ean}",
        "product_link_sel": "a[class*='product-tile'], a[href*='/product/']",
        "price_sel": "[class*='price__value'], [data-testid='price-value']",
        "name_sel": "h1[class*='product-title'], h1[class*='name']",
        "stock_sel": "[class*='availability'], [class*='stock-status']",
    },
}


def scrape_site(ean: str, site_key: str, config: dict, session: requests.Session) -> dict:
    """Scrape a single site for a given EAN."""
    result = {
        "site": site_key,
        "ean": ean,
        "name": None,
        "price_str": None,
        "price_float": None,
        "in_stock": None,
        "url": None,
        "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "status": "success",
        "error": None,
    }

    try:
        search_url = config["search_url"].format(ean=ean)
        resp = session.get(search_url, headers=HEADERS, timeout=15)
        resp.raise_for_status()

        soup = BeautifulSoup(resp.text, "html.parser")
        link = soup.select_one(config["product_link_sel"])

        if not link:
            result["status"] = "not_found"
            result["error"] = f"No results on {site_key}"
            return result

        href = link.get("href", "")
        base = "https://" + site_key
        result["url"] = href if href.startswith("http") else base + href

        # Fetch product page
        time.sleep(0.4)
        prod = session.get(result["url"], headers=HEADERS, timeout=15)
        prod.raise_for_status()
        psoup = BeautifulSoup(prod.text, "html.parser")

        # Name
        name_el = psoup.select_one(config["name_sel"])
        result["name"] = name_el.get_text(strip=True) if name_el else None

        # Price
        price_el = psoup.select_one(config["price_sel"])
        if price_el:
            raw = price_el.get_text(strip=True)
            result["price_str"] = raw
            # Parse Dutch price format: € 12,99 → 12.99
            m = re.search(r"[\d]+[,.][\d]{2}", raw)
            if m:
                result["price_float"] = float(m.group().replace(",", "."))

        # Stock
        stock_el = psoup.select_one(config["stock_sel"])
        if stock_el:
            stock_text = stock_el.get_text(strip=True).lower()
            result["in_stock"] = "voorraad" in stock_text or "beschikbaar" in stock_text

    except Exception as e:
        result["status"] = "error"
        result["error"] = str(e)

    return result


def compare_prices(ean: str) -> pd.DataFrame:
    """Compare prices for one EAN across all configured sites."""
    session = requests.Session()
    results = []

    print(f"\nComparing prices for EAN: {ean}")
    print("-" * 50)

    for site_key, config in SITES.items():
        print(f"  Checking {site_key}...", end=" ")
        r = scrape_site(ean, site_key, config, session)
        results.append(r)
        if r["status"] == "success" and r["price_float"]:
            print(f"€ {r['price_float']:.2f}")
        else:
            print(r.get("error") or r["status"])
        time.sleep(0.8)

    df = pd.DataFrame(results)

    # Find best price
    valid = df[df["price_float"].notna()]
    if not valid.empty:
        best_idx = valid["price_float"].idxmin()
        print(f"\n  ✓ Best price: {df.loc[best_idx, 'site']} — € {df.loc[best_idx, 'price_float']:.2f}")

    return df


def export_comparison(df: pd.DataFrame, output_path: str = None) -> str:
    """Export comparison results to Excel with conditional formatting."""
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"price_comparison_{ts}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary sheet
        summary_cols = ["ean", "site", "name", "price_str", "price_float", "in_stock", "url", "scraped_at", "status"]
        df[summary_cols].to_excel(writer, index=False, sheet_name="Price Comparison")

        ws = writer.sheets["Price Comparison"]

        from openpyxl.styles import PatternFill, Font
        green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        bold = Font(bold=True)

        # Highlight best price per EAN
        for ean_val in df["ean"].unique():
            subset = df[df["ean"] == ean_val]
            valid = subset[subset["price_float"].notna()]
            if not valid.empty:
                best_idx = valid["price_float"].idxmin()
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    cell_ean = ws.cell(row=row_idx, column=1).value
                    cell_site = ws.cell(row=row_idx, column=2).value
                    if cell_ean == ean_val:
                        if cell_site == df.loc[best_idx, "site"]:
                            for cell in ws[row_idx]:
                                cell.fill = green_fill
                                cell.font = bold

        # Auto-size
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        ws.freeze_panes = "A2"

    return output_path


def main():
    import sys
    eans = sys.argv[1:] if len(sys.argv) > 1 else [
        input("Enter EAN code: ").strip()
    ]

    all_results = []
    for ean in eans:
        df = compare_prices(ean.strip())
        all_results.append(df)

    final_df = pd.concat(all_results, ignore_index=True)
    out = export_comparison(final_df)
    print(f"\n✓ Exported to: {out}")


if __name__ == "__main__":
    main()